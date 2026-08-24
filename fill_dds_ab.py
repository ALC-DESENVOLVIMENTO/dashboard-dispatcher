from datetime import datetime
import json
import re
import unicodedata
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl

src_path = Path(r"C:\Users\Wesley\Downloads\Rotas 1 a 31 Julho.xlsx")
target_path = Path(r"C:\Users\Wesley\Downloads\Rotas 1 a 31 Julho DDS.xlsx")
out_path = Path(r"C:\Users\Wesley\Documents\Dev Alc\Projetos\bonificacao-dispatcher\outputs\thread-01\Rotas 1 a 31 Julho DDS preenchida.xlsx")
log_path = out_path.with_suffix('.json')

def norm_name(value):
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', text).upper().strip()

def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value or '')[:10]

source_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
source_ws = source_wb['Métricas']
source_by_name_date = defaultdict(list)
for row in source_ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None or row[3] is None:
        continue
    key = (iso_date(row[0]), norm_name(row[3]))
    source_by_name_date[key].append((str(row[1]).strip() if row[1] is not None else '', row[11]))
source_wb.close()

target_wb = openpyxl.load_workbook(target_path)
target_ws = target_wb['Geral']
filled = 0
unmatched = []
ambiguous = []
for row_num in range(2, target_ws.max_row + 1):
    date_value = target_ws.cell(row_num, 2).value
    name_value = target_ws.cell(row_num, 4).value
    route_value = target_ws.cell(row_num, 10).value
    if date_value is None or name_value is None:
        continue
    key = (iso_date(date_value), norm_name(name_value))
    candidates = source_by_name_date.get(key, [])
    chosen = None
    if len(candidates) == 1:
        chosen = candidates[0][1]
    elif len(candidates) > 1:
        route_id = str(route_value).strip() if route_value is not None else ''
        route_matches = [value for rid, value in candidates if rid == route_id]
        if len(route_matches) == 1:
            chosen = route_matches[0]
        else:
            ambiguous.append({'row': row_num, 'date': iso_date(date_value), 'motorista': name_value, 'rota': route_id, 'candidates': candidates})
    else:
        unmatched.append({'row': row_num, 'date': iso_date(date_value), 'motorista': name_value, 'rota': route_value})
    if chosen is not None:
        target_ws.cell(row_num, 28).value = chosen
        filled += 1

target_wb.save(out_path)
log = {
    'output': str(out_path),
    'sheet': 'Geral',
    'column': 'AB',
    'filled': filled,
    'unmatched': len(unmatched),
    'ambiguous': len(ambiguous),
    'unmatched_sample': unmatched[:20],
    'ambiguous': ambiguous,
}
log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(json.dumps({k: v for k, v in log.items() if k not in ('unmatched_sample', 'ambiguous')}, ensure_ascii=False))
print('UNMATCHED_SAMPLE', json.dumps(unmatched[:10], ensure_ascii=False, default=str))
print('AMBIGUOUS_SAMPLE', json.dumps(ambiguous[:10], ensure_ascii=False, default=str))
